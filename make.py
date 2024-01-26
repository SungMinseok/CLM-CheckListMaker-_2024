import pandas as pd
from openpyxl import load_workbook

def process_data_template(data_file, template_file, output_file, key_column):
    # Read data and template files
    data_df = pd.read_excel(data_file)
    template_wb = load_workbook(template_file)
    
    # Iterate through each row in the data
    for index, row in data_df.iterrows():
        data_dict = row.to_dict()
        
        # Iterate through each sheet in the template
        for sheet_name in template_wb.sheetnames:
            template_ws = template_wb[sheet_name]

            # Check if the key column exists in the template sheet
            if key_column in template_ws['A']:
                # Find the row index corresponding to the key value in the template
                template_key_row = template_ws['A'].index(key_column)
                template_row = template_ws[template_key_row + 1]

                # Replace placeholders with data values
                for col_num, col_value in enumerate(template_row, start=1):
                    if col_value is not None and '{' in col_value and '}' in col_value:
                        col_name = col_value.replace('{', '').replace('}', '').strip()
                        if col_name in data_dict:
                            template_ws.cell(row=template_key_row + index + 2, column=col_num).value = data_dict[col_name]

    # Save the modified template to a new Excel file
    template_wb.save(output_file)

# Example usage
process_data_template('data.xlsx', 'template.xlsx', 'output.xlsx', 'CashShop ID')

