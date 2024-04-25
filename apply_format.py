import openpyxl
import time
def get_format_dict(cell):
    #cell = ws.unmerge_cells(get_cell.coordinate)[0][0]  
    return {
        "font": cell.font,
        "alignment": cell.alignment,
        "fill": cell.fill,
        "border": cell.border,
    #"width": ws.column_dimensions[cell.column_letter].width,
       # "height": ws.row_dimensions[cell.row].height,
    }


def is_merged_cell(ws, row, column):
    # 병합된 셀 범위 내에 해당 셀이 포함되는지 확인
    for merged_range in ws.merged_cells.ranges:
        if (row >= merged_range.min_row and row <= merged_range.max_row and
            column >= merged_range.min_col and column <= merged_range.max_col):
            return True
    return False

def check_and_merge_cells(target_ws, start_row, start_column, end_row, end_column):
    # 병합하려는 범위 내에 이미 병합된 셀이 있는지 확인
    for row in range(start_row, end_row + 1):
        for col in range(start_column, end_column + 1):
            if is_merged_cell(target_ws, row, col):
                return  # 이미 병합된 셀이 있으면 병합하지 않음

    # 병합 실행
    target_ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row+1, end_column=end_column)

def apply_template(target_xlsx_file, template_xlsx_file, template_sheet_name, output_name):
    template_wb = openpyxl.load_workbook(template_xlsx_file)
    target_wb = openpyxl.load_workbook(target_xlsx_file)
    
    template_ws = template_wb[template_sheet_name]
    target_ws = target_wb.active

    print(f'{template_ws.max_row=}')
    print(f'{target_ws.max_row=}')

    total_count = int( (target_ws.max_row-1) / (template_ws.max_row-1) )

    # 최초 1회만 셀 서식 저장
    format_dict = {}
    for template_row in template_ws.iter_rows(min_row=2, max_row=template_ws.max_row, max_col=template_ws.max_column):
        for template_cell in template_row:
            format_dict[template_cell.coordinate] = get_format_dict(template_cell)

    # 서식 적용
    for i in range(total_count):
        for row_target, row_template in zip(target_ws.iter_rows(min_row=i*(template_ws.max_row-1)+1, max_row=target_ws.max_row, max_col=template_ws.max_column), template_ws.iter_rows(min_row=1, max_row=template_ws.max_row, max_col=template_ws.max_column)):
            for target_cell, template_cell in zip(row_target, row_template):
                target_format = format_dict.get(template_cell.coordinate, {})
                print(target_cell.value)
                if "font" in target_format:
                    target_cell.font = openpyxl.styles.Font(**target_format["font"].__dict__)

                if "alignment" in target_format:
                    target_cell.alignment = openpyxl.styles.Alignment(**target_format["alignment"].__dict__)

                if "fill" in target_format:
                    target_cell.fill = openpyxl.styles.PatternFill(**target_format["fill"].__dict__)

                if "border" in target_format:
                    target_cell.border = openpyxl.styles.Border(**target_format["border"].__dict__)

                if "width" in target_format:
                    target_cell.width = openpyxl.styles.Width(**target_format["width"].__dict__)

        #for row_target, row_template in zip(target_ws.iter_rows(min_row=i*(template_ws.max_row-1)+1, max_row=target_ws.max_row, max_col=template_ws.max_column), template_ws.iter_rows(min_row=1, max_row=template_ws.max_row, max_col=template_ws.max_column)):
        #    for target_cell, template_cell in zip(row_target, row_template):
                # if target_cell.value is None:
                #     # 현재 셀 값이 None인 경우 위로 이동하면서 값이 None이 아닌 곳까지 찾아서 병합
                #     above_cell = target_cell
                #     while above_cell.value is None and above_cell.row > 1:
                #         above_cell = target_ws.cell(row=above_cell.row - 1, column=above_cell.column)

                #     if above_cell.value is not None:
                #         # 병합 대상이 발견되면 병합
                #         target_ws.merge_cells(start_row=above_cell.row, start_column=above_cell.column, end_row=target_cell.row, end_column=target_cell.column)
    
                #240425                
                if target_cell.value is None:
                    # 현재 셀 값이 None인 경우 위로 이동하면서 값이 None이 아닌 곳까지 찾아서 병합
                    above_cell = target_cell
                    while above_cell.value is None and above_cell.row > 1:
                        above_cell = target_ws.cell(row=above_cell.row - 1, column=above_cell.column)

                    if above_cell.value is not None:
                        # 병합 대상이 발견되면 병합
                        check_and_merge_cells(target_ws, above_cell.row, above_cell.column, target_cell.row, target_cell.column)
    
    
    # # 서식 적용
    # for target_row in target_ws.iter_rows(min_row=2, max_row=target_ws.max_row, max_col=template_ws.max_column):
    #     for target_cell, template_cell in zip(target_row, template_ws.iter_cols(min_row=1, max_row=template_ws.max_row, max_col=template_ws.max_column)):
    #         target_format = format_dict.get(template_cell.coordinate, {})
    #         target_cell.font = target_format.get("font", openpyxl.styles.Font())
    #         target_cell.alignment = target_format.get("alignment", openpyxl.styles.Alignment())
    #         target_cell.fill = target_format.get("fill", openpyxl.styles.PatternFill())
    #         target_cell.border = target_format.get("border", openpyxl.styles.Border())
    # 템플릿 문서의 열 길이와 타겟 문서의 열 길이 동일하게 조정

    # for template_col, target_col in zip(template_ws.iter_cols(), target_ws.iter_cols()):
    #     target_col[0].column_dimensions = template_col[0].column_dimensions

    #target_ws.column_dimensions['a'].width = template_ws.column_dimensions['a'].width 






    # # 각 열에 대해 반복
    # for col in target_ws.iter_cols():
    #     start_row = 0
    #     start_value = None
    #     for cell in col:
    #         if cell.value is not None:
    #             if start_row == 0:
    #                 # 처음으로 값이 나타나는 경우 시작 행 및 값 설정
    #                 start_row = cell.row
    #                 start_value = cell.value
    #             elif cell.value != start_value:
    #                 # 이전 값과 현재 값이 다른 경우 병합
    #                 merge_target_cell = f"{col[0].column_letter}{start_row}:{col[0].column_letter}{cell.row - 1}"
    #                 target_ws.merge_cells(merge_target_cell)
    #                 # 다음 병합을 위해 시작 행 및 값 업데이트
    #                 start_value = cell.value
    #                 start_row = cell.row
    #     # 마지막으로 남은 범위에 대해 병합
    #     if start_row != 0 and start_row != col[-1].row:
    #         merge_target_cell = f"{col[0].column_letter}{start_row}:{col[0].column_letter}{col[-1].row}"
    #         target_ws.merge_cells(merge_target_cell)






























    for col_letter in template_ws.iter_cols(min_row=1, max_row=1):
        col_letter = col_letter[0].column_letter
        target_ws.column_dimensions[col_letter].width = template_ws.column_dimensions[col_letter].width
    target_wb.save(output_name)


# 예제 사용법
#apply_template("1.target_xlsx_file.xlsx", "2.template_xlsx_file.xlsx")
if __name__ == "__main__" :
    target_file = fr'd:\파이썬결과물저장소\CLM\result\Cashshop_20240208_125510.xlsx'
    template_file = fr'd:\파이썬결과물저장소\CLM\template.xlsx'
    sheet_name = 'Cashshop'
    output_name = fr'd:\파이썬결과물저장소\CLM\test_{time.strftime("%Y%m%d_%H%M%S")}_format.xlsx'
    apply_template(target_file,template_file,sheet_name,output_name)